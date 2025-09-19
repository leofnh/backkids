import openpyxl
from ..models import Produto, Cliente
from .produtos import get_products
from decimal import Decimal
from datetime import datetime
from django.db.models import Q

def import_products(file):
    wb = openpyxl.load_workbook(file)
    sheet = wb.active
    last_row = sheet.max_row
    print('Esta no bd certo!')
    for row in sheet.iter_rows(min_row=2, max_row=last_row, min_col=3, 
                               max_col=11):
        # Tratamento seguro dos valores das células
        marca = row[0].value.strip() if row[0].value else ""
        marca_sem_strip = row[0].value
        tamanho = row[1].value.strip() if row[1].value else ""
        preco = row[2].value
        data = row[3].value
        qtde = row[4].value if row[4].value is not None else 0
        codigo = row[5].value.strip() if row[5].value else ""
        ref = row[6].value.strip() if row[6].value else ""
        custo = row[7].value if row[7].value is not None else 0
        
        # Pula linha se código estiver vazio (obrigatório)
        if not codigo:            
            continue
        
        # Função para fazer parse da data com múltiplos formatos
        def parse_data(data_valor):
            if not data_valor:
                return datetime.now()
            
            # Se já é um objeto datetime, retorna como está
            if isinstance(data_valor, datetime):
                return data_valor
            
            # Se é string, tenta diferentes formatos
            if isinstance(data_valor, str):
                data_str = data_valor.strip()
                formatos_data = [
                    '%d/%m/%Y %H:%M:%S',  # Formato original com hora
                    '%d/%m/%Y',           # Formato apenas data
                    '%Y-%m-%d %H:%M:%S',  # Formato ISO com hora
                    '%Y-%m-%d',           # Formato ISO apenas data
                    '%d-%m-%Y',           # Formato com traços
                    '%d/%m/%y',           # Ano com 2 dígitos
                ]
                
                for formato in formatos_data:
                    try:
                        return datetime.strptime(data_str, formato)
                    except ValueError:
                        continue
                
                # Se nenhum formato funcionou, usa data atual
                return datetime.now()
            
            # Para outros tipos, usa data atual
            return datetime.now()
        
        data = parse_data(data)
        
        # Tratamento seguro do preço
        def parse_valor_monetario(valor, nome_campo="valor"):
            if valor is None:
                return 0.0
            
            if isinstance(valor, str):
                # Remove símbolos de moeda e normaliza
                valor_limpo = valor.replace('R$', '').replace(',', '.').strip()
                try:
                    return float(Decimal(valor_limpo))
                except (ValueError, TypeError):
                    print(f"Aviso: Não foi possível converter {nome_campo} '{valor}', usando 0.0")
                    return 0.0
            elif isinstance(valor, (int, float)):
                return float(valor)
            else:
                print(f"Aviso: {nome_campo} de tipo desconhecido '{type(valor)}', usando 0.0")
                return 0.0
        
        preco = parse_valor_monetario(preco, "preço")
        custo = parse_valor_monetario(custo, "custo")
        
        try:
            ja_existe = Produto.objects.filter(Q(codigo=codigo) |
                                                Q(ref=ref) |
                                                Q(marca=marca) |
                                                Q(marca=marca_sem_strip))
            if ja_existe.exists():
                # Atualiza produto existente
                ja_existe.update(
                    preco=preco,
                    # estoque=qtde,
                    # custo=custo,
                    # marca=marca,
                    # tamanho=tamanho,
                    # ref=ref,
                    # cadastro=data
                )                
            else:            
                # Cria novo produto
                Produto.objects.create(
                    preco=preco,
                    tamanho=tamanho,
                    marca=marca,
                    ref=ref,
                    custo=custo,
                    estoque=qtde,
                    codigo=codigo,
                    cadastro=data
                )
                
                
        except Exception as e:
            print(f"Erro ao processar produto {codigo}: {str(e)}")
            continue
    
    dados = get_products()
    dados_json = list(dados.values('codigo', 'estoque', 'marca', 'custo', 
                                       'preco', 'ref', 'tamanho'))
     
    
    return dados_json

def import_cliente(file):
    """
    Importa clientes de arquivo Excel com tratamento robusto de dados
    """
    wb = openpyxl.load_workbook(file)
    sheet = wb.active
    last_row = sheet.max_row
    
    # Função para fazer parse da data (reutilizada)
    def parse_data_cliente(data_valor):
        if not data_valor:
            return datetime.now()
        
        if isinstance(data_valor, datetime):
            return data_valor
        
        if isinstance(data_valor, str):
            data_str = data_valor.strip()
            formatos_data = [
                '%d/%m/%Y %H:%M:%S',
                '%d/%m/%Y',
                '%Y-%m-%d %H:%M:%S',
                '%Y-%m-%d',
                '%d-%m-%Y',
                '%d/%m/%y',
            ]
            
            for formato in formatos_data:
                try:
                    return datetime.strptime(data_str, formato)
                except ValueError:
                    continue
            
            print(f"Aviso: Não foi possível fazer parse da data do cliente '{data_str}', usando data atual")
            return datetime.now()
        
        return datetime.now()
    
    for row in sheet.iter_rows(min_row=2, max_row=last_row, min_col=2, max_col=7):
        # Tratamento seguro dos valores das células
        nome = row[0].value.strip() if row[0].value else ""
        email = row[1].value.strip() if row[1].value else ""
        telefone = row[2].value.strip() if row[2].value else ""
        data = row[3].value
        cidade = row[4].value.strip() if row[4].value else ""
        estado = row[5].value.strip() if row[5].value else ""
        
        # Pula linha se email estiver vazio (obrigatório para cliente)
        if not email:
            print(f"Aviso: Linha de cliente pulada - email vazio")
            continue
        
        # Parse da data
        data_cadastro = parse_data_cliente(data)
        
        try:
            ja_existe = Cliente.objects.filter(email=email)
            if ja_existe.exists():
                # Atualiza cliente existente
                ja_existe.update(
                    nome=nome,
                    telefone=telefone,
                    cidade=cidade,
                    estado=estado,
                    cadastro=data_cadastro
                )
                print(f"Cliente atualizado: {email}")
            else:            
                # Cria novo cliente
                Cliente.objects.create(
                    nome=nome,
                    email=email,
                    telefone=telefone,
                    cidade=cidade,
                    estado=estado,
                    cadastro=data_cadastro
                )
                print(f"Cliente criado: {email}")
                
        except Exception as e:
            print(f"Erro ao processar cliente {email}: {str(e)}")
            continue

      

    
    